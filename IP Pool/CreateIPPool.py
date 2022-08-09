import requests
import urllib3
import json
import getpass
import colorama
from colorama import init
init()
from colorama import Fore,Back,Style
from openpyxl import load_workbook
import pandas as pd

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)


def sort_excel(file):
    xl = pd.ExcelFile(file)
    df = xl.parse("Sheet1")
    df = df.sort_values(by="Fortinet IP")
    output_file_name = 'sorted_fCreateIPPool.xlsx'
    writer = pd.ExcelWriter(output_file_name)
    df.to_excel(writer,sheet_name='Sheet1',index=False)
    writer.save()
    return output_file_name


def create_ip_pool(fgtip,vdom,token,ip_pool_name, ip_pool_type, start_ip, end_ip):

    if (vdom == None):
        vdom = "root"
        
    if (ip_pool_type == None):
        ip_pool_type = "overload"
    
    if (end_ip == None):
        end_ip = start_ip
    
    if (start_ip == None):
        start_ip = end_ip

    if (ip_pool_name == None):
        if ((end_ip == None) or (end_ip == start_ip)):
            ip_pool_name = start_ip
        else:
            ip_pool_name = str(start_ip) + "-" + str(end_ip)
    

    ip_pool_link = '/firewall/ippool'
    ip_pool_url = "https://" + fgtip + "/api/v2/cmdb" + ip_pool_link + "?access_token=" + token + "&vdom=" + vdom
    ip_pool_data = {
        "name": ip_pool_name,
        "type": ip_pool_type,
        "startip": start_ip,
        "endip": end_ip      
    }
    

    ip_pool_body = json.dumps(ip_pool_data)
    create_result = requests.post(url=ip_pool_url, data=ip_pool_body, verify=False)
    return create_result.status_code, ip_pool_name
    

if __name__ == '__main__':

    file_sorted_excel = sort_excel("fCreateIPPool.xlsx")
    wb = load_workbook(file_sorted_excel, data_only=True)
    sheet = wb["Sheet1"]

    temp = 0
    
    for row in sheet.iter_rows(min_row=2,min_col=1,max_col=8,values_only=True):
        fgtip = row[0]
        vdom = row[1]
        ip_pool_name = row[2]
        ip_pool_type = row[3]
        start_ip = row[4]
        end_ip = row[5]

        
        if fgtip != temp and fgtip != None:
            text = "Enter token of fortigate " + str(fgtip) + " : "
            token = input(text)
        if fgtip == None:
            print (Fore.CYAN + "The Excel file was ended."+ Style.RESET_ALL)
            break
        else:
            pass
        
        result_code ,ip_pool_name = create_ip_pool(fgtip,vdom,token,ip_pool_name, ip_pool_type, start_ip, end_ip)
        temp = fgtip
        
        
        if result_code!=200:
            print (Fore.RED + "Create IP Pool %s" %ip_pool_name ," on host %s" %fgtip,"is %s"%result_code+ Style.RESET_ALL)
        else:
            if vdom ==None:
                vdom ="root"
            #print (Fore.GREEN + "Create IP Pool %s" %ip_pool_name ," on host %s" %fgtip,"is successfully"+ Style.RESET_ALL)
            print (Fore.GREEN + "Create IP Pool %s on host %s on VDOM %s is successfully" %(ip_pool_name, fgtip, vdom) + Style.RESET_ALL)
            
        
